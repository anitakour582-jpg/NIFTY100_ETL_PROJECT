"""
Sprint 3 - Day 20

Peer Comparison Excel Export
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from peer import prepare_peer_data, calculate_peer_percentiles


OUTPUT_FILE = "output/peer_comparison.xlsx"


# ----------------------------------------
# Export Excel
# ----------------------------------------

def export_peer_comparison():

    os.makedirs(
        "output",
        exist_ok=True
    )


    df = prepare_peer_data()

    percentiles = calculate_peer_percentiles()


    merged = df.merge(
        percentiles,
        on=[
            "company_id",
            "peer_group_name",
            "year"
        ],
        how="left"
    )


    peer_groups = (
        merged["peer_group_name"]
        .dropna()
        .unique()
    )


    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:


        for group in peer_groups:

            sheet = merged[
                merged["peer_group_name"] == group
            ].copy()


            sheet = sheet.sort_values(
                "company_name"
            )


            sheet.to_excel(
                writer,
                sheet_name=group[:31],
                index=False
            )


    apply_formatting()


    print(
        "Peer comparison exported successfully."
    )

    print(
        OUTPUT_FILE
    )


# ----------------------------------------
# Formatting
# ----------------------------------------

def apply_formatting():

    workbook = load_workbook(
        OUTPUT_FILE
    )


    green = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )


    yellow = PatternFill(
        start_color="FFEB9C",
        end_color="FFEB9C",
        fill_type="solid"
    )


    red = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )


    for worksheet in workbook:

        percentile_columns = []

        for cell in worksheet[1]:

            if "percentile" in str(cell.value).lower():

                percentile_columns.append(
                    cell.column
                )


        for column in percentile_columns:

            for row in range(
                2,
                worksheet.max_row + 1
            ):

                cell = worksheet.cell(
                    row=row,
                    column=column
                )


                try:

                    value = float(
                        cell.value
                    )


                    if value >= 75:

                        cell.fill = green

                    elif value >= 25:

                        cell.fill = yellow

                    else:

                        cell.fill = red


                except:

                    pass


    workbook.save(
        OUTPUT_FILE
    )



if __name__ == "__main__":

    export_peer_comparison()