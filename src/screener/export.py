"""
Sprint 3 - Day 17

Composite Score + Excel Export
"""

import pandas as pd
import numpy as np

from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from engine import load_financial_data
from presets import run_all_presets


# ----------------------------------------
# Winsorisation + Normalisation
# ----------------------------------------

def normalise(series, reverse=False):

    s = pd.to_numeric(series, errors="coerce").fillna(0)

    p10 = s.quantile(0.10)
    p90 = s.quantile(0.90)

    s = s.clip(lower=p10, upper=p90)

    minimum = s.min()
    maximum = s.max()

    if maximum == minimum:
        return pd.Series(50, index=s.index)

    score = (s - minimum) / (maximum - minimum) * 100

    if reverse:
        score = 100 - score

    return score


# ----------------------------------------
# Composite Score
# ----------------------------------------

def calculate_composite_score(df):

    data = df.copy()

    data["roe_score"] = normalise(
        data["return_on_equity_pct"]
    )

    data["roce_score"] = normalise(
        data["return_on_equity_pct"]
    )

    data["npm_score"] = normalise(
        data["net_profit_margin_pct"]
    )

    data["fcf_score"] = normalise(
        data["free_cash_flow_cr"]
    )

    data["cfo_score"] = normalise(
        data["cash_from_operations_cr"]
    )

    if "revenue_cagr_5y" in data.columns:
        data["rev_score"] = normalise(
            data["revenue_cagr_5y"]
        )
    else:
        data["rev_score"] = 50

    if "pat_cagr_5y" in data.columns:
        data["pat_score"] = normalise(
            data["pat_cagr_5y"]
        )
    else:
        data["pat_score"] = 50

    data["de_score"] = normalise(
        data["debt_to_equity"],
        reverse=True
    )

    data["icr_score"] = normalise(
        data["interest_coverage"]
    )

    return data


def add_composite_quality_score(df):

    data = calculate_composite_score(df)

    data["composite_quality_score"] = (

        data["roe_score"] * 0.15 +

        data["roce_score"] * 0.10 +

        data["npm_score"] * 0.10 +

        data["fcf_score"] * 0.15 +

        data["cfo_score"] * 0.10 +

        data["rev_score"] * 0.10 +

        data["pat_score"] * 0.10 +

        data["de_score"] * 0.10 +

        data["icr_score"] * 0.10

    )

    if "broad_sector" in data.columns:

        data["sector_relative_score"] = (
            data.groupby("broad_sector")[
                "composite_quality_score"
            ].rank(pct=True) * 100
        )

    else:

        data["sector_relative_score"] = (
            data["composite_quality_score"]
            .rank(pct=True) * 100
        )

    return data.sort_values(
        "composite_quality_score",
        ascending=False
    )# ----------------------------------------
# Excel Export
# ----------------------------------------

def export_screener_results(df):

    df = add_composite_quality_score(df)

    preset_results = run_all_presets(df)

    output_file = "output/screener_output.xlsx"

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        for preset_name, preset_df in preset_results.items():

            preset_df = preset_df.sort_values(
                "composite_quality_score",
                ascending=False
            )

            preset_df.to_excel(
                writer,
                sheet_name=preset_name[:31],
                index=False
            )

            worksheet = writer.sheets[preset_name[:31]]

            # Freeze header
            worksheet.freeze_panes = "A2"

            # Header formatting
            header_fill = PatternFill(
                start_color="1F4E78",
                end_color="1F4E78",
                fill_type="solid"
            )

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

            # Auto-width columns
            for column_cells in worksheet.columns:

                length = max(
                    len(str(cell.value))
                    if cell.value is not None else 0
                    for cell in column_cells
                )

                worksheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = min(length + 3, 40)

            # Conditional formatting
            if "composite_quality_score" in preset_df.columns:

                score_column = (
                    preset_df.columns.get_loc(
                        "composite_quality_score"
                    ) + 1
                )

                green = PatternFill(
                    start_color="C6EFCE",
                    end_color="C6EFCE",
                    fill_type="solid"
                )

                yellow = PatternFill(
                    start_color="FFEB9C",
                    end_color="FFEB9C",
                    fill_type="solid"
                )

                red = PatternFill(
                    start_color="F4CCCC",
                    end_color="F4CCCC",
                    fill_type="solid"
                )

                for row in range(2, worksheet.max_row + 1):

                    cell = worksheet.cell(
                        row=row,
                        column=score_column
                    )

                    try:
                        score = float(cell.value)

                        if score >= 80:
                            cell.fill = green
                        elif score >= 60:
                            cell.fill = yellow
                        else:
                            cell.fill = red

                    except (TypeError, ValueError):
                        pass

    print("Excel exported successfully.")
    print(output_file)

    return output_file # ----------------------------------------
# Main
# ----------------------------------------

if __name__ == "__main__":

    df = load_financial_data()

    export_screener_results(df)